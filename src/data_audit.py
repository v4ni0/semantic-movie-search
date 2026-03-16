import os
import pandas as pd
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image

mpl.rcParams["text.usetex"] = False
mpl.rcParams["text.parse_math"] = False

class DataAudit:
    IMAGES_DIR = os.path.join(os.path.dirname(__file__), "images")
    def __init__(self, data: pd.DataFrame, name = None):
        self.data = data
        self.name = name

        if not os.path.exists(self.IMAGES_DIR):
            os.mkdir(self.IMAGES_DIR)

    def add_hist_sheets(self, wb: Workbook):

        for col in self.data.columns:
            sheet_name = col
            ws = wb.create_sheet(title=sheet_name)
            fig, ax = plt.subplots(figsize=(7, 4), dpi=160)
            img_path = os.path.join(self.IMAGES_DIR, f"{col}_hist.png")
            counts = self.data[col].value_counts(dropna=False).sort_index()
            temp = pd.DataFrame({"Value": counts.index, "Counts": counts.values})
            ws["A1"] = "Value"
            ws["B1"] = "Counts"
            row_idx = 2
            for _, row in temp.iterrows():
                ws.cell(row=row_idx, column=1, value=row["Value"])
                ws.cell(row=row_idx, column=2, value=row["Counts"])
                row_idx += 1
            if pd.api.types.is_numeric_dtype(self.data[col]):
                if len(counts) > 10:
                    sns.histplot(self.data[col], ax=ax, edgecolor="black", bins=30)
                    ax.set_title(f"Histogram of {col}")
                    ax.set_xlabel(col)
                    ax.set_ylabel("Frequency")
                else:
                    safe_index = [str(v).replace("$", r"\$") for v in counts.index]
                    sns.barplot(x=safe_index, y=counts.values, ax=ax, edgecolor="black")
                    ax.set_title(f"Bar Plot of {col}")
                    ax.set_xlabel(col)
                    ax.set_ylabel("Counts")
            else:
                if len(counts) <=50:
                    safe_index = [str(v).replace("$", r"\$") for v in counts.index]
                    sns.barplot(x=safe_index, y=counts.values, ax=ax, edgecolor="black")
                    ax.set_title(f"Bar Plot of {col}")
                    ax.set_xlabel(col)
                    ax.set_ylabel("Counts")

            fig.savefig(img_path, bbox_inches="tight")
            img = Image(img_path)
            ws.add_image(img, "E1")
            fig.tight_layout()

    def statistical_summary(self):
        desc = pd.DataFrame(index=self.data.columns)                  
        desc = desc.join(self.data.describe().T, how="left")          
        desc["mode"] = self.data.mode().iloc[0]
        desc["unique_count"] = self.data.nunique()
        desc["missing_count"] = self.data.isnull().sum()
        desc["missing_pct"] = self.data.isnull().mean() * 100
        return desc
    
    def create_data_audit(self):
        desc = self.statistical_summary()
        wb = Workbook()
        ws = wb.active
        ws.append([""] + desc.columns.tolist())
        for feature, row in desc.iterrows():
            ws.append([feature] + list(row.values))

        self.add_hist_sheets(wb)
        if self.name:
            wb.save(self.name)

        print(desc)

        